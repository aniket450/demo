import sys
import os
import smtplib, ssl
import re
import openpyxl


def write_excel(scr_no, commit_msg, owner_name, review_name):
    New_path = os.getcwd()
    New_path = New_path + '\\SCR_Tracker.xlsx'
    workbook = openpyxl.load_workbook(New_path)
    sheet = workbook.active
    strt_row = sheet.max_row + 1
    li = list(scr_no)
    scrlist = []
    msglist = []
    for col in sheet['A']:
        scrlist.append(col.value)
    scrlist.remove('SCR')
    print(scrlist)

    if scr_no in scrlist:
        row_no = scrlist.index(scr_no) + 1
        status_col = sheet.cell(row=row_no + 1, column=3)
        status_col.value = "CLOSED"
        dec_col = sheet.cell(row=row_no + 1, column=4)
        dec_col.value = commit_msg
        owner_col = sheet.cell(row=row_no + 1, column=5)
        owner_col.value = owner_name
        reviewer_col = sheet.cell(row=row_no + 1, column=6)
        reviewer_col.value = review_name
        workbook.save(New_path)
        print("Closed ", scr_no)
    else:
        print("Add SCR in excel sheet")


comment = os.environ['GERRIT_CHANGE_SUBJECT']    # SCR_001- Fixed issue review:murugan
owner_name = os.environ['GERRIT_CHANGE_OWNER_NAME']


split_desc = comment.split('-')
scr_no = split_desc[0]
desc = split_desc[1]
split_desc = desc.split(':')
review_name = split_desc[1]
scr_desc = split_desc[0]
if scr_desc.endswith('Review'):
     scr_desc = scr_desc.replace('Review', '')
else:
    scr_desc = scr_desc.replace('review', '')

write_excel(scr_no, scr_desc, owner_name, review_name)

